import win32com.client
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas.io.excel._openpyxl
import csv
import sys
from datetime import datetime
import os


def fileOK(fileName):
    try:
        f = open(fileName, "rb")
        f.close()
        return True
    except:
        return False


def decrypt(fileName, password, newFileName):
    if os.path.exists(newFileName):
        os.remove(newFileName)
    xlApp = win32com.client.Dispatch("Excel.Application")
    if os.path.exists(newFileName):
        print("**** Error: File %s could not be found correct the file name in Files.xlsx" % (fileName))
        return False
    if debug:
        print("Decrypt file:%s Version:%s" % (fileName, xlApp.Version))
    try:
        wb = xlApp.Workbooks.Open(fileName, False, True,
                                  None, password, password, False)
    except:
        print("**** Error: File %s might be open in excel - please close it or correct the file name in Files.xlsx" % (fileName))
        return False
    wb.SaveAs(newFileName, 51, "")
    wb.Close()
    return True


def addTo(wb, sheet, master, first):
    ws = wb[sheet]
    masterWs = master[sheet]
    for i in range(1, ws.max_row):
        c2 = str(ws.cell(i+1, 2).value)
        c3 = str(ws.cell(i+1, 3).value)
        if(c2 != 'None' and debug):
            print(c2, c3)
    row1 = True
    count = 0
    for r in ws.rows:
        if(str(r[1].value) != "None") and (first or (not row1)):
            nr = []
            for c in r:
                s = str(c.value)
                if s == 'None':
                    s = ''
                nr.append(s)
            if debug:
                print(sheet, nr)
            masterWs.append(nr)
            count += 1
        row1 = False
    return count


def appendTo(fileName, master, first):
    if debug:
        print("Append %s file" % (fileName))
    wb = load_workbook(filename=fileName)
    count1 = addTo(wb, 'Active', master, first)
    count2 = addTo(wb, "RTW", master, first)
    return (count1, count2)


def getFiles(fileName):
    results = []
    try:
        wb = load_workbook(filename=fileName)
        ws = wb['Sheet1']
    except:
        print("**** Error file %s could not be opened or does not have correct format" % (fileName))
        return []
    for i in range(1, ws.max_row):
        file = str(ws.cell(i+1, 2).value)
        password = str(ws.cell(i+1, 3).value)
        if password == 'None':
            password = ''
        results.append([file, password])
    wb.close()
    return results


def addTestData(masterWs, sheetName):
    data = [
        ['Apples', 10000, 5000, 8000, 6000],
        ['Pears',   2000, 3000, 4000, 5000],
        ['Bananas', 6000, 6000, 6500, 6000],
        ['Oranges',  500,  300,  200,  700],
    ]

    masterWs = masterWb[sheetName]
    # add column headings. NB. these must be strings
    masterWs.append([sheetName, "2011", "2012", "2013", "2014"])
    for row in data:
        masterWs.append(row)
    tab = Table(displayName=sheetName, ref="A1:E5")
    masterWs.add_table(tab)


debug = False
root = "c:\\FunStuff\\colleen\\"
# root = os.getcwd() + "\\"
if debug:
    print("root path", root, sys.path[0], os.getcwd())
res = getFiles(root + "Files.xlsx")
if(len(res) == 0):
    exit()
if(debug):
    print(res)

now = datetime.now()
mast = "Master" + now.strftime("-%Y-%m-%d") + ".xlsx"
masterWb = Workbook()
masterWb.create_sheet("Active")
masterWb.create_sheet("RTW")

# Test adding data
# addTestData(masterWb, "Active")
# addTestData(masterWb, "RTW")

first = True
rows = 0
for r in res:
    if decrypt(root + r[0], r[1], root + 'test.xlsx'):
        counts = appendTo(root + 'test.xlsx', masterWb, first)
        print("Added %s to %s %d Active and %d RTW rows" %
              (r[0], mast, counts[0], counts[1]))
        rows += counts[0] + counts[1]
        first = False
        os.remove(root + 'test.xlsx')

# For some silly reason a sheet called Sheet is added remove it
#sh = masterWb.get_sheet_by_name("Sheet")
# masterWb.remove_sheet(sh)
masterWs = masterWb["Sheet"]
masterWb.remove(masterWs)

# Save the master file
#if os.path.exists(root + "Master.xlsx"):
if os.path.exists(root + mast):
    os.remove(root + mast)
try:
    masterWb.save(root + mast)
    print("Complete -- writing the %s file with %d new rows" % (mast, rows))
except:
    print("The file " + mast + " can not be saved -- it might be open in excel -- please close it")